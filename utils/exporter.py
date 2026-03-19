import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from utils.logger import app_logger

class Exporter:
    def __init__(self, filename: str = "trade_logs.xlsx"):
        self.filename = filename
        self._init_workbook()

    def _init_workbook(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Anomalies"
            # Из статьи: Колонки для анализа спреда
            ws.append(["Average", "Futures", "Percentage", "Time", "Symbol", "Signal", "Strength"])
            wb.save(self.filename)
            app_logger.info(f"💾 Создан файл аналитики: {self.filename}")

    def log_anomaly(self, avg: float, futures: float, pct: float, symbol: str, signal: str, strength: int):
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            times = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([avg, futures, pct, times, symbol, signal, strength])
            wb.save(self.filename)
        except Exception as e:
            app_logger.error(f"Ошибка записи в Excel: {e}")

# Global instance
exporter = Exporter("data_analysis.xlsx")
